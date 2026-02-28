"""
Cyclic Dependency Finder
========================
Reads an Excel file (all sheets) and detects circular dependencies.

Usage:
    python find_cycles.py myfile.xlsx
"""

import sys
from collections import defaultdict

import openpyxl


# ---------------------------------------------------------------------------
# Interactive prompts
# ---------------------------------------------------------------------------

def ask_choice(prompt, options):
    """Show a numbered menu and return the user's choice."""
    print()
    print(prompt)
    for i, option in enumerate(options, 1):
        print(f"  {i}. {option}")

    while True:
        try:
            choice = int(input(f"Enter number (1-{len(options)}): "))
            if 1 <= choice <= len(options):
                return options[choice - 1]
        except (ValueError, EOFError):
            pass
        print(f"  Please enter a number between 1 and {len(options)}.")


def collect_columns(workbook):
    """
    Scan all sheets, gather every unique column name, and let the user
    pick the source and target columns interactively.
    """
    all_columns = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        header_row = next(sheet.iter_rows(min_row=1, max_row=1), None)
        if header_row is None:
            continue
        for cell in header_row:
            if cell.value is not None:
                name = str(cell.value).strip()
                if name and name not in all_columns:
                    all_columns.append(name)

    if len(all_columns) < 2:
        print("ERROR: The Excel file must have at least 2 columns.")
        sys.exit(1)

    print()
    print("=" * 60)
    print("  Columns found in your Excel file:")
    print("=" * 60)
    for i, col in enumerate(all_columns, 1):
        print(f"  {i}. {col}")

    source_col = ask_choice(
        "Which column contains the ITEM NAME (e.g. Task, Module)?",
        all_columns,
    )

    remaining = [c for c in all_columns if c != source_col]
    target_col = ask_choice(
        "Which column contains the DEPENDENCY (e.g. Depends On, Requires)?",
        remaining,
    )

    return source_col, target_col


def ask_separator():
    """Ask how multiple dependencies are separated in a single cell."""
    print()
    print("If one cell can list multiple dependencies, how are they separated?")
    print("  1. Comma        (e.g. Task A, Task B)")
    print("  2. Semicolon    (e.g. Task A; Task B)")
    print("  3. Pipe         (e.g. Task A | Task B)")
    print("  4. New line     (e.g. each on its own line inside the cell)")
    print("  5. Only one dependency per cell (no separator needed)")

    while True:
        try:
            choice = int(input("Enter number (1-5): "))
            if choice in (1, 2, 3, 4, 5):
                break
        except (ValueError, EOFError):
            pass
        print("  Please enter a number between 1 and 5.")

    mapping = {1: ",", 2: ";", 3: "|", 4: "\n", 5: None}
    return mapping[choice]


# ---------------------------------------------------------------------------
# Graph helpers
# ---------------------------------------------------------------------------

def build_graph(workbook, source_col, target_col, separator):
    """
    Read every sheet in the workbook and build a directed graph.

    Each row creates edges:  source_value  -->  target_value

    If a target cell contains multiple values separated by `separator`,
    each value becomes a separate edge.
    """
    graph = defaultdict(set)
    edge_origins = defaultdict(list)

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        headers = [
            str(cell.value).strip() if cell.value is not None else ""
            for cell in next(sheet.iter_rows(min_row=1, max_row=1))
        ]

        headers_lower = [h.lower() for h in headers]
        src_lower = source_col.lower()
        tgt_lower = target_col.lower()

        if src_lower not in headers_lower or tgt_lower not in headers_lower:
            continue

        src_idx = headers_lower.index(src_lower)
        tgt_idx = headers_lower.index(tgt_lower)

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[src_idx] is None or row[tgt_idx] is None:
                continue

            source_val = str(row[src_idx]).strip()
            if not source_val:
                continue

            raw_targets = str(row[tgt_idx]).strip()
            if not raw_targets:
                continue

            if separator:
                targets = [t.strip() for t in raw_targets.split(separator) if t.strip()]
            else:
                targets = [raw_targets]

            for target_val in targets:
                graph[source_val].add(target_val)
                edge_origins[(source_val, target_val)].append(sheet_name)
                if target_val not in graph:
                    graph[target_val] = set()

    return graph, edge_origins


def find_all_cycles(graph):
    """Find all cycles in a directed graph using DFS."""
    WHITE, GRAY, BLACK = 0, 1, 2
    cycles = []

    def dfs(node, color, path, path_set):
        color[node] = GRAY
        path.append(node)
        path_set.add(node)

        for neighbour in graph.get(node, []):
            if color[neighbour] == GRAY and neighbour in path_set:
                cycle_start = path.index(neighbour)
                cycles.append(path[cycle_start:])
            elif color[neighbour] == WHITE:
                dfs(neighbour, color, path, path_set)

        path.pop()
        path_set.discard(node)
        color[node] = BLACK

    color = {node: WHITE for node in graph}
    for node in graph:
        if color[node] == WHITE:
            dfs(node, color, [], set())

    return cycles


def deduplicate_cycles(cycles):
    """Remove duplicate cycles that are just rotations of each other."""
    seen = set()
    unique = []
    for cycle in cycles:
        min_idx = cycle.index(min(cycle))
        rotated = tuple(cycle[min_idx:] + cycle[:min_idx])
        if rotated not in seen:
            seen.add(rotated)
            unique.append(cycle)
    return unique


# ---------------------------------------------------------------------------
# Display
# ---------------------------------------------------------------------------

def print_cycles(cycles, edge_origins):
    """Pretty-print each cycle with sheet origin information."""
    if not cycles:
        print()
        print("=== RESULT: No cyclic dependencies found! ===")
        print()
        return

    print()
    print(f"=== RESULT: Found {len(cycles)} cyclic dependency(ies)! ===")
    print()

    for i, cycle in enumerate(cycles, 1):
        chain = " -> ".join(cycle) + " -> " + cycle[0]
        print(f"  Cycle {i}:  {chain}")

        for j in range(len(cycle)):
            src = cycle[j]
            tgt = cycle[(j + 1) % len(cycle)]
            sheets = edge_origins.get((src, tgt), ["unknown"])
            print(f"            {src} -> {tgt}  (found in sheet: {', '.join(sheets)})")
        print()


def print_summary(workbook, graph, source_col, target_col):
    """Print a summary of what was read."""
    print()
    print("=" * 60)
    print("  Cyclic Dependency Finder - Summary")
    print("=" * 60)
    print(f"  Sheets scanned  : {len(workbook.sheetnames)}")
    print(f"  Source column    : {source_col}")
    print(f"  Target column    : {target_col}")
    print(f"  Unique items     : {len(graph)}")
    total_edges = sum(len(v) for v in graph.values())
    print(f"  Total links      : {total_edges}")
    print("=" * 60)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    # --- Get Excel file path ---
    if len(sys.argv) >= 2:
        excel_path = sys.argv[1]
    else:
        print()
        print("=" * 60)
        print("  Cyclic Dependency Finder")
        print("=" * 60)
        excel_path = input("Enter the path to your Excel file (.xlsx): ").strip()
        # Remove surrounding quotes if user copy-pasted a path with quotes
        excel_path = excel_path.strip('"').strip("'")

    if not excel_path:
        print("ERROR: No file path provided.")
        sys.exit(1)

    # --- Load workbook ---
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    except FileNotFoundError:
        print(f"ERROR: File not found: {excel_path}")
        sys.exit(1)
    except Exception as e:
        print(f"ERROR: Could not open Excel file: {e}")
        sys.exit(1)

    print(f"\n  Loaded: {excel_path}")
    print(f"  Sheets: {', '.join(wb.sheetnames)}")

    # --- Interactive column selection ---
    source_col, target_col = collect_columns(wb)
    separator = ask_separator()

    # --- Build graph ---
    graph, edge_origins = build_graph(wb, source_col, target_col, separator)

    if not graph:
        print(f"\nERROR: No data found with columns '{source_col}' and '{target_col}'.")
        sys.exit(1)

    print_summary(wb, graph, source_col, target_col)

    # --- Detect cycles ---
    raw_cycles = find_all_cycles(graph)
    cycles = deduplicate_cycles(raw_cycles)

    print_cycles(cycles, edge_origins)

    wb.close()

    # Pause so the window stays open if user double-clicked the script
    input("Press Enter to exit...")


if __name__ == "__main__":
    main()
