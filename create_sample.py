"""Creates a sample Excel file with cyclic dependencies for testing."""

import openpyxl

wb = openpyxl.Workbook()

# Sheet 1 - has a cycle: A -> B -> C -> A
ws1 = wb.active
ws1.title = "ProjectTasks"
ws1.append(["Task", "Depends On"])
ws1.append(["Task A", "Task B"])
ws1.append(["Task B", "Task C"])
ws1.append(["Task C", "Task A"])
ws1.append(["Task D", "Task B"])

# Sheet 2 - has another cycle: X -> Y -> Z -> X
ws2 = wb.create_sheet("Modules")
ws2.append(["Task", "Depends On"])
ws2.append(["Module X", "Module Y"])
ws2.append(["Module Y", "Module Z"])
ws2.append(["Module Z", "Module X"])

# Sheet 3 - no cycles here
ws3 = wb.create_sheet("NoCycles")
ws3.append(["Task", "Depends On"])
ws3.append(["Step 1", "Step 2"])
ws3.append(["Step 2", "Step 3"])

wb.save("sample_dependencies.xlsx")
print("Created: sample_dependencies.xlsx")
print("Run:  python find_cycles.py sample_dependencies.xlsx --source \"Task\" --target \"Depends On\"")
